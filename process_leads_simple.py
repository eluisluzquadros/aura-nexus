#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AURA NEXUS - Processamento Simplificado de Leads
"""

import io
import sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import asyncio
import argparse
from pathlib import Path
import logging
from datetime import datetime
import pandas as pd
from dotenv import load_dotenv

# Adiciona src ao path
sys.path.insert(0, str(Path(__file__).parent))

from src.infrastructure.spreadsheet_adapter import SpreadsheetAdapter
from src.infrastructure.cache_system import SmartMultiLevelCache
from src.infrastructure.checkpoint_manager import CheckpointManager
from src.core.api_manager import APIManager
from src.core.lead_processor import LeadProcessor
from src.core.multi_llm_consensus import MultiLLMConsensus

# Carrega variáveis de ambiente
load_dotenv()

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger("AURA_NEXUS")


# Mapas de features por modo
FEATURE_MODES = {
    'basic': ['google_details', 'contact_extraction'],
    'full': ['google_details', 'google_cse', 'web_scraping', 'social_scraping', 'contact_extraction'],
    'premium': ['google_details', 'google_cse', 'web_scraping', 'social_scraping', 'contact_extraction', 'ai_analysis']
}


async def process_leads(args):
    """Processa leads com o modo especificado"""
    
    print("\n=== AURA NEXUS - PROCESSAMENTO DE LEADS ===")
    print(f"📍 Arquivo: {args.input}")
    print(f"🔧 Modo: {args.mode}")
    print(f"📦 Batch size: {args.batch_size}")
    print(f"⚡ Workers: {args.max_workers}\n")
    
    # 1. Carregar planilha
    print("1️⃣ Carregando planilha...")
    adapter = SpreadsheetAdapter()
    df = adapter.adapt_spreadsheet(args.input)
    print(f"   ✅ {len(df)} leads carregados")
    
    # Aplicar limite se especificado
    if args.limit:
        df = df.head(args.limit)
        print(f"   📍 Limitado a {len(df)} leads")
    
    # 2. Inicializar componentes
    print("\n2️⃣ Inicializando componentes...")
    
    # Cache
    cache = SmartMultiLevelCache(Path("data/cache"))
    await cache.initialize()
    print("   ✅ Cache inicializado")
    
    # Checkpoint Manager
    checkpoint = CheckpointManager()
    session = checkpoint.create_session(f"process_{datetime.now():%Y%m%d_%H%M%S}")
    print(f"   ✅ Sessão de checkpoint: {session}")
    
    # API Manager
    api_manager = APIManager()
    await api_manager.initialize()
    apis = api_manager.get_available_apis()
    print(f"   ✅ APIs disponíveis: {apis}")
    
    # Lead Processor
    processor = LeadProcessor(api_manager, cache)
    await processor.initialize()
    print("   ✅ Lead Processor pronto")
    
    # Multi-LLM Consensus (se modo premium)
    consensus = None
    if args.mode == 'premium':
        consensus = MultiLLMConsensus(api_manager)
        print(f"   ✅ Multi-LLM Consensus: {consensus.available_llms}")
    
    # 3. Processar leads
    print(f"\n3️⃣ Processando {len(df)} leads...")
    
    features = FEATURE_MODES.get(args.mode, FEATURE_MODES['basic'])
    print(f"   📋 Features: {features}")
    
    results = []
    errors = []
    
    # Processar em batches
    for i in range(0, len(df), args.batch_size):
        batch = df.iloc[i:i+args.batch_size]
        batch_num = i // args.batch_size + 1
        
        print(f"\n   🔄 Batch {batch_num}/{(len(df) + args.batch_size - 1) // args.batch_size}")
        
        # Processar batch em paralelo
        tasks = []
        for idx, row in batch.iterrows():
            lead_data = row.to_dict()
            
            # Se já tem Google ID e não deve forçar
            if lead_data.get('gdr_ja_enriquecido_google') and not args.force_all:
                lead_data['skip_google_api'] = True
            
            # Criar task
            task = process_single_lead(processor, lead_data, features, idx+1)
            tasks.append(task)
        
        # Executar batch
        batch_results = await asyncio.gather(*tasks[:args.max_workers], return_exceptions=True)
        
        # Processar resultados
        for result in batch_results:
            if isinstance(result, Exception):
                errors.append(str(result))
                logger.error(f"❌ Erro: {result}")
            else:
                results.append(result)
        
        # Salvar checkpoint a cada batch
        checkpoint_data = {
            'processed': len(results),
            'errors': len(errors),
            'last_batch': batch_num
        }
        await checkpoint.save_checkpoint(checkpoint_data, f"batch_{batch_num}")
        
        # Aguardar entre batches para não sobrecarregar APIs
        if i + args.batch_size < len(df):
            await asyncio.sleep(2)
    
    # 4. Salvar resultados
    print(f"\n4️⃣ Salvando resultados...")
    
    if results:
        # Criar DataFrame com resultados
        results_df = pd.DataFrame(results)
        
        # Ordenar colunas para melhor organização
        results_df = organize_columns(results_df)
        
        # Definir nome do arquivo de saída
        output_file = args.output
        if not output_file:
            input_path = Path(args.input)
            output_file = input_path.parent / f"{input_path.stem}_enriched_{args.mode}.xlsx"
        
        # Salvar Excel com formatação melhorada
        save_enhanced_excel(results_df, output_file, args.mode)
        print(f"   ✅ Resultados salvos em: {output_file}")
        
        # Estatísticas detalhadas
        print_detailed_statistics(results_df, errors)
        
        # Relatório de colunas
        print_column_report(results_df)
    
    # Fechar recursos
    await processor.close()
    await api_manager.close()
    
    print("\n✅ Processamento concluído!")


async def process_single_lead(processor, lead_data, features, lead_num):
    """Processa um único lead"""
    try:
        logger.info(f"   → Lead {lead_num}: {lead_data['nome_empresa']}")
        
        # Processar lead
        result = await processor.process_lead(lead_data, features)
        
        # Log de conclusão (after flattening, keys are different)
        features_ok = result.get('gdr_features_executadas', '')
        total_errors = result.get('gdr_total_erros', 0)
        
        if total_errors > 0:
            logger.warning(f"   ⚠️ Lead {lead_num} com {total_errors} erros")
        else:
            feature_count = result.get('gdr_total_features_executadas', 0)
            if isinstance(features_ok, str):
                feature_count = len(features_ok.split(';')) if features_ok else 0
            logger.info(f"   ✅ Lead {lead_num} processado: {feature_count} features")
        
        return result
        
    except Exception as e:
        logger.error(f"   ❌ Erro no lead {lead_num}: {e}")
        raise


def organize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    Organiza colunas do DataFrame para melhor visualização
    """
    # Definir ordem preferencial das colunas
    priority_columns = [
        # Informações básicas
        'nome_empresa', 'nome_fantasia', 'cnpj', 'cidade', 'estado',
        
        # Google Maps
        'google_maps_place_id', 'google_maps_nome', 'google_maps_endereco',
        'google_maps_telefone', 'google_maps_website', 'google_maps_rating',
        'google_maps_total_avaliacoes', 'google_maps_status',
        
        # Website Info
        'website_info_url', 'website_info_titulo', 'website_info_descricao',
        'website_info_emails', 'website_info_telefones',
        
        # Contatos consolidados
        'contatos_emails', 'contatos_telefones', 'contatos_websites',
        'contatos_total_contatos',
        
        # Redes sociais
        'redes_sociais_links', 'website_info_redes_sociais_facebook',
        'website_info_redes_sociais_instagram', 'website_info_redes_sociais_linkedin',
        
        # AI Analysis
        'ai_analysis_status', 'ai_analysis_score', 'ai_analysis_analise',
        'ai_analysis_pontos_fortes', 'ai_analysis_oportunidades',
        'ai_analysis_recomendacao',
        
        # Traceability
        'gdr_processamento_inicio', 'gdr_processamento_fim',
        'gdr_features_executadas', 'gdr_total_features_executadas',
        'gdr_total_erros', 'gdr_taxa_sucesso'
    ]
    
    # Obter colunas existentes
    existing_columns = [col for col in priority_columns if col in df.columns]
    remaining_columns = [col for col in df.columns if col not in existing_columns]
    
    # Reorganizar DataFrame
    ordered_columns = existing_columns + sorted(remaining_columns)
    return df[ordered_columns]


def save_enhanced_excel(df: pd.DataFrame, output_file: str, mode: str):
    """
    Salva DataFrame em Excel com formatação melhorada
    """
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Salvar dados principais
        df.to_excel(writer, sheet_name='Leads_Enriched', index=False)
        
        # Criar planilha de resumo
        create_summary_sheet(df, writer, mode)
        
        # Criar planilha de mapeamento de colunas
        create_column_mapping_sheet(df, writer)
        
        # Aplicar formatação
        format_excel_sheets(writer, df)


def create_summary_sheet(df: pd.DataFrame, writer: pd.ExcelWriter, mode: str):
    """
    Cria planilha de resumo com estatísticas
    """
    summary_data = {
        'Métrica': [
            'Total de Leads',
            'Leads com Google Maps',
            'Leads com Website',
            'Leads com Contatos',
            'Leads com Análise IA',
            'Taxa Sucesso Média',
            'Modo de Processamento',
            'Data/Hora Processamento'
        ],
        'Valor': [
            len(df),
            sum(1 for x in df.get('google_maps_place_id', []) if pd.notna(x)),
            sum(1 for x in df.get('website_info_url', []) if pd.notna(x)),
            sum(1 for x in df.get('contatos_total_contatos', []) if pd.notna(x) and x > 0),
            sum(1 for x in df.get('ai_analysis_score', []) if pd.notna(x) and x > 0),
            f"{df.get('gdr_taxa_sucesso', [0]).mean():.1f}%",
            mode.upper(),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ]
    }
    
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, sheet_name='Resumo', index=False)


def create_column_mapping_sheet(df: pd.DataFrame, writer: pd.ExcelWriter):
    """
    Cria planilha explicando o significado das colunas
    """
    column_descriptions = {
        # Básicas
        'nome_empresa': 'Nome da empresa',
        'cnpj': 'CNPJ da empresa',
        'cidade': 'Cidade',
        'estado': 'Estado',
        
        # Google Maps
        'google_maps_place_id': 'ID único do Google Maps',
        'google_maps_nome': 'Nome no Google Maps',
        'google_maps_endereco': 'Endereço completo do Google Maps',
        'google_maps_telefone': 'Telefone do Google Maps',
        'google_maps_website': 'Website do Google Maps',
        'google_maps_rating': 'Avaliação no Google Maps (1-5)',
        'google_maps_total_avaliacoes': 'Total de avaliações no Google Maps',
        'google_maps_status': 'Status do Google Maps',
        
        # Website
        'website_info_url': 'URL do website',
        'website_info_titulo': 'Título da página',
        'website_info_descricao': 'Meta descrição',
        'website_info_emails': 'E-mails encontrados no site',
        'website_info_telefones': 'Telefones encontrados no site',
        
        # Contatos
        'contatos_emails': 'E-mails consolidados',
        'contatos_telefones': 'Telefones consolidados',
        'contatos_websites': 'Websites consolidados',
        'contatos_total_contatos': 'Total de contatos encontrados',
        
        # IA
        'ai_analysis_score': 'Score de potencial de negócio (0-100)',
        'ai_analysis_analise': 'Análise detalhada da IA',
        'ai_analysis_pontos_fortes': 'Pontos fortes identificados',
        'ai_analysis_oportunidades': 'Oportunidades identificadas',
        'ai_analysis_recomendacao': 'Recomendação da IA',
        
        # Rastreabilidade
        'gdr_processamento_inicio': 'Início do processamento',
        'gdr_processamento_fim': 'Fim do processamento',
        'gdr_features_executadas': 'Features executadas',
        'gdr_total_features_executadas': 'Total de features executadas',
        'gdr_total_erros': 'Total de erros',
        'gdr_taxa_sucesso': 'Taxa de sucesso (%)',
        'gdr_google_maps_status': 'Status Google Maps',
        'gdr_website_scraping_status': 'Status Website Scraping',
        'gdr_ai_analysis_status': 'Status Análise IA'
    }
    
    mapping_data = []
    for col in df.columns:
        description = column_descriptions.get(col, 'Coluna gerada automaticamente')
        data_type = str(df[col].dtype)
        sample_value = str(df[col].iloc[0]) if len(df) > 0 and pd.notna(df[col].iloc[0]) else 'N/A'
        
        mapping_data.append({
            'Coluna': col,
            'Descrição': description,
            'Tipo': data_type,
            'Exemplo': sample_value[:100]  # Limitar exemplo
        })
    
    mapping_df = pd.DataFrame(mapping_data)
    mapping_df.to_excel(writer, sheet_name='Mapeamento_Colunas', index=False)


def format_excel_sheets(writer: pd.ExcelWriter, df: pd.DataFrame):
    """
    Aplica formatação nas planilhas Excel
    """
    from openpyxl.styles import Font, Fill, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    # Acessar workbook
    workbook = writer.book
    
    # Formatar planilha principal
    if 'Leads_Enriched' in workbook.sheetnames:
        ws = workbook['Leads_Enriched']
        
        # Formatar cabeçalhos
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Máximo 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width


def print_detailed_statistics(df: pd.DataFrame, errors: list):
    """
    Imprime estatísticas detalhadas do processamento
    """
    print(f"\n📊 Estatísticas Detalhadas:")
    print(f"   ✅ Processados: {len(df)}")
    print(f"   ❌ Erros: {len(errors)}")
    print(f"   📈 Taxa de sucesso: {len(df)/(len(df)+len(errors))*100:.1f}%")
    
    # Estatísticas por feature
    if 'gdr_google_maps_status' in df.columns:
        google_success = sum(1 for x in df['gdr_google_maps_status'] if 'sucesso' in str(x))
        print(f"   🗺️ Google Maps: {google_success}/{len(df)} ({google_success/len(df)*100:.1f}%)")
    
    if 'gdr_website_scraping_status' in df.columns:
        website_success = sum(1 for x in df['gdr_website_scraping_status'] if 'sucesso' in str(x))
        print(f"   🌐 Website Scraping: {website_success}/{len(df)} ({website_success/len(df)*100:.1f}%)")
    
    if 'contatos_total_contatos' in df.columns:
        contacts_found = sum(1 for x in df['contatos_total_contatos'] if pd.notna(x) and x > 0)
        print(f"   📞 Contatos Encontrados: {contacts_found}/{len(df)} ({contacts_found/len(df)*100:.1f}%)")
    
    if 'ai_analysis_score' in df.columns:
        ai_analyzed = sum(1 for x in df['ai_analysis_score'] if pd.notna(x) and x > 0)
        avg_score = df['ai_analysis_score'].mean() if ai_analyzed > 0 else 0
        print(f"   🤖 Análise IA: {ai_analyzed}/{len(df)} - Score médio: {avg_score:.1f}")


def print_column_report(df: pd.DataFrame):
    """
    Imprime relatório das colunas geradas
    """
    print(f"\n📋 Relatório de Colunas:")
    print(f"   📊 Total de colunas: {len(df.columns)}")
    
    # Categorizar colunas
    categories = {
        'Básicas': [col for col in df.columns if any(x in col for x in ['nome_empresa', 'cnpj', 'cidade', 'estado', 'endereco'])],
        'Google Maps': [col for col in df.columns if col.startswith('google_maps_')],
        'Website': [col for col in df.columns if col.startswith('website_info_')],
        'Contatos': [col for col in df.columns if col.startswith('contatos_')],
        'Redes Sociais': [col for col in df.columns if 'social' in col or any(x in col for x in ['facebook', 'instagram', 'linkedin'])],
        'IA Analysis': [col for col in df.columns if col.startswith('ai_analysis_')],
        'Rastreabilidade': [col for col in df.columns if col.startswith('gdr_')]
    }
    
    for category, columns in categories.items():
        if columns:
            print(f"   • {category}: {len(columns)} colunas")
            for col in columns[:3]:  # Mostrar apenas as 3 primeiras
                print(f"     - {col}")
            if len(columns) > 3:
                print(f"     ... e mais {len(columns)-3} colunas")


def main():
    """Função principal"""
    parser = argparse.ArgumentParser(description="AURA NEXUS - Processamento de Leads")
    
    parser.add_argument("--input", "-i", required=True, help="Arquivo Excel de entrada")
    parser.add_argument("--output", "-o", help="Arquivo Excel de saída")
    parser.add_argument("--mode", "-m", choices=['basic', 'full', 'premium'], default='basic', help="Modo de processamento")
    parser.add_argument("--batch-size", type=int, default=5, help="Tamanho do batch")
    parser.add_argument("--max-workers", type=int, default=3, help="Workers paralelos")
    parser.add_argument("--force-all", action="store_true", help="Força reprocessamento")
    parser.add_argument("--limit", type=int, help="Limita número de leads")
    
    args = parser.parse_args()
    
    # Executar processamento
    asyncio.run(process_leads(args))


if __name__ == "__main__":
    main()